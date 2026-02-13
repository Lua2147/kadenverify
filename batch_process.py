#!/usr/bin/env python3
"""KadenVerify Batch Processor — end-to-end contact file processing.

Ingests xlsx/csv contact files, finds missing emails via the enrichment waterfall,
verifies all emails, iteratively squeezes risky/unknown, and exports merged
contact data + verification results to xlsx.

Usage:
  python batch_process.py run ./contacts/           # Process all xlsx/csv in directory
  python batch_process.py run file.xlsx             # Process a single file
  python batch_process.py run ./contacts/ -o out.xlsx --squeeze 10
  python batch_process.py export ./output/state.csv -o results.xlsx
  python batch_process.py stats ./output/state.csv
"""

import asyncio
import csv
import json
import logging
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiohttp
import click

logger = logging.getLogger("kadenverify.batch")

# ── Defaults ────────────────────────────────────────────────────────────────

DEFAULT_API_URL = os.environ.get("KADENVERIFY_API_URL", "http://198.23.249.137:8025")
DEFAULT_API_KEY = os.environ.get("KADENVERIFY_API_KEY", "")  # pragma: allowlist secret

FIND_BATCH_SIZE = 50
FIND_CONCURRENCY = 6
VERIFY_BATCH_SIZE = 50
VERIFY_CONCURRENCY = 10
SQUEEZE_COOLDOWN = 20
MAX_SQUEEZE_ITERATIONS = 14

# ── Column mapping (PitchBook xlsx conventions) ─────────────────────────────

NAME_COLUMNS = ["Full Name", "full_name", "Name", "name", "Contact Name", "Contact Full Name", "member_full_name"]
FIRST_NAME_COLUMNS = ["First Name", "first_name", "FirstName", "Contact First Name", "member_name_first"]
LAST_NAME_COLUMNS = ["Last Name", "last_name", "LastName", "Contact Last Name", "member_name_last"]
EMAIL_COLUMNS = ["Email Address", "Email", "email", "email_address", "Work Email", "Contact Primary E-mail Address", "E-mail Address"]
COMPANY_COLUMNS = ["Primary Company", "Company", "company", "Organization", "Company Name", "company_name"]
WEBSITE_COLUMNS = ["Primary Company Website", "Website", "website", "Domain", "Company Website", "Company Website", "domain"]
POSITION_COLUMNS = ["Primary Title", "Title", "Position", "Job Title", "title", "position", "job_title", "Contact Title"]
PHONE_COLUMNS = ["Phone", "phone", "Phone Number", "Direct Phone", "Contact Primary Phone Number"]
LINKEDIN_COLUMNS = ["LinkedIn", "linkedin", "LinkedIn URL", "LinkedIn Profile"]
LOCATION_COLUMNS = ["Location", "location", "City", "Geography"]
PROFILE_COLUMNS = ["Profile URL", "PitchBook Profile", "URL"]


def _find_column(headers: list[str], candidates: list[str]) -> Optional[str]:
    """Find the first matching column name from candidates."""
    header_set = {h.strip() for h in headers}
    for c in candidates:
        if c in header_set:
            return c
    # Case-insensitive fallback
    lower_map = {h.lower().strip(): h for h in headers}
    for c in candidates:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None


def _guess_domain(company: str) -> str:
    """Guess domain from company name when website isn't available."""
    if not company:
        return ""
    c = company.lower().strip()
    c = re.sub(
        r"\b(inc|llc|ltd|corp|co|group|holdings|partners|capital|management|"
        r"ventures|advisors|consulting|international|services|solutions|"
        r"technologies|global|the)\b",
        "",
        c,
    )
    c = re.sub(r"[^a-z0-9]", "", c).strip()
    return f"{c}.com" if c else ""


# ── File Loading ────────────────────────────────────────────────────────────


def load_contacts_from_file(filepath: Path) -> list[dict]:
    """Load contacts from a single xlsx or csv file, normalizing column names."""
    contacts = []

    if filepath.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        # Try to find the best sheet: prefer "Contacts" by name, then most contact-like columns
        ws = wb.active
        if len(wb.sheetnames) > 1:
            # First: look for a sheet literally named "Contacts"
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() == "contacts":
                    ws = wb[sheet_name]
                    break
            else:
                # Fallback: pick the sheet with the most contact-related columns
                best_score, best_ws = -1, None
                for sheet_name in wb.sheetnames:
                    candidate = wb[sheet_name]
                    first_row = next(candidate.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        hdrs = [str(h).strip().lower() for h in first_row if h]
                        score = sum(1 for h in hdrs for kw in ["email", "e-mail", "first name", "last name", "full name"] if kw in h)
                        if score > best_score:
                            best_score, best_ws = score, candidate
                if best_ws is not None:
                    ws = best_ws
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                d[h] = str(val).strip() if val is not None else ""
            contacts.append(d)

    elif filepath.suffix.lower() == ".csv":
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                contacts.append({k: (v or "").strip() for k, v in row.items()})

    else:
        return []

    if not contacts:
        return []

    # Normalize to standard fields
    headers = list(contacts[0].keys())
    name_col = _find_column(headers, NAME_COLUMNS)
    first_col = _find_column(headers, FIRST_NAME_COLUMNS)
    last_col = _find_column(headers, LAST_NAME_COLUMNS)
    email_col = _find_column(headers, EMAIL_COLUMNS)
    company_col = _find_column(headers, COMPANY_COLUMNS)
    website_col = _find_column(headers, WEBSITE_COLUMNS)
    position_col = _find_column(headers, POSITION_COLUMNS)
    phone_col = _find_column(headers, PHONE_COLUMNS)
    linkedin_col = _find_column(headers, LINKEDIN_COLUMNS)
    location_col = _find_column(headers, LOCATION_COLUMNS)
    profile_col = _find_column(headers, PROFILE_COLUMNS)

    normalized = []
    for row in contacts:
        # Build full name
        full_name = row.get(name_col, "") if name_col else ""
        if not full_name and first_col and last_col:
            full_name = f"{row.get(first_col, '')} {row.get(last_col, '')}".strip()
        if not full_name:
            continue

        # Parse first/last from full name
        parts = full_name.split()
        first_name = row.get(first_col, "") if first_col else (parts[0] if parts else "")
        last_name = row.get(last_col, "") if last_col else (parts[-1] if len(parts) >= 2 else "")

        # Domain from website or company guess
        website = row.get(website_col, "") if website_col else ""
        company = row.get(company_col, "") if company_col else ""
        domain = ""
        if website:
            domain = website.replace("https://", "").replace("http://", "").replace("www.", "").strip("/").split("/")[0].split(":")[0]
        if not domain and company:
            domain = _guess_domain(company)

        email = row.get(email_col, "") if email_col else ""
        email = email.lower().strip() if email and "@" in email else ""

        normalized.append({
            "first_name": first_name,
            "last_name": last_name,
            "full_name": full_name,
            "email": email,
            "company": company,
            "domain": domain,
            "position": row.get(position_col, "") if position_col else "",
            "website": website,
            "phone": row.get(phone_col, "") if phone_col else "",
            "linkedin": row.get(linkedin_col, "") if linkedin_col else "",
            "location": row.get(location_col, "") if location_col else "",
            "profile_url": row.get(profile_col, "") if profile_col else "",
            "source_file": filepath.name,
        })

    return normalized


def load_contacts(input_path: Path) -> list[dict]:
    """Load contacts from a file or directory of files."""
    contacts = []
    if input_path.is_dir():
        files = sorted(
            f for f in input_path.iterdir()
            if f.suffix.lower() in (".xlsx", ".xls", ".csv") and not f.name.startswith("~")
        )
        for f in files:
            batch = load_contacts_from_file(f)
            click.echo(f"  {f.name}: {len(batch)} contacts")
            contacts.extend(batch)
    elif input_path.is_file():
        contacts = load_contacts_from_file(input_path)
        click.echo(f"  {input_path.name}: {len(contacts)} contacts")
    else:
        click.echo(f"Error: {input_path} not found")
        sys.exit(1)
    return contacts


# ── API Calls ───────────────────────────────────────────────────────────────


async def _api_find_batch(
    session: aiohttp.ClientSession,
    contacts: list[dict],
    api_url: str,
    headers: dict,
    label: str = "FIND",
) -> list[dict]:
    """Find emails for a batch of contacts via the API."""
    if not contacts:
        return []

    batches = [contacts[i : i + FIND_BATCH_SIZE] for i in range(0, len(contacts), FIND_BATCH_SIZE)]
    sem = asyncio.Semaphore(FIND_CONCURRENCY)
    all_results = []
    completed = [0]

    async def do_batch(batch, batch_num):
        async with sem:
            payload = []
            for c in batch:
                if c.get("first_name") and c.get("last_name") and c.get("domain"):
                    payload.append({
                        "first_name": c["first_name"],
                        "last_name": c["last_name"],
                        "domain": c["domain"],
                        "company_name": c.get("company", ""),
                    })
            if not payload:
                return []
            try:
                async with session.post(
                    f"{api_url}/find-email/batch",
                    json={"contacts": payload},
                    headers=headers,
                    timeout=aiohttp.ClientTimeout(total=600),
                ) as resp:
                    if resp.status == 200:
                        results = await resp.json()
                        results = results if isinstance(results, list) else results.get("results", [])
                        found = sum(1 for r in results if (r.get("email") or "") and float(r.get("confidence", 0) or 0) > 0)
                        completed[0] += len(batch)
                        click.echo(f"    [{label}] Batch {batch_num}/{len(batches)}: {found}/{len(results)} found ({completed[0]}/{len(contacts)})")
                        await asyncio.sleep(2)
                        return results
                    else:
                        text = await resp.text()
                        click.echo(f"    [{label}] Batch {batch_num} error: {resp.status} {text[:100]}")
                        return []
            except Exception as e:
                click.echo(f"    [{label}] Batch {batch_num} error: {e}")
                return []

    tasks = [do_batch(batch, i + 1) for i, batch in enumerate(batches)]
    for coro in asyncio.as_completed(tasks):
        all_results.extend(await coro)
    return all_results


async def _api_verify_batch(
    session: aiohttp.ClientSession,
    emails: list[str],
    api_url: str,
    headers: dict,
    label: str = "VERIFY",
    batch_size: int = VERIFY_BATCH_SIZE,
    concurrency: int = VERIFY_CONCURRENCY,
) -> list[dict]:
    """Verify a batch of emails via the API."""
    if not emails:
        return []

    batches = [emails[i : i + batch_size] for i in range(0, len(emails), batch_size)]
    sem = asyncio.Semaphore(concurrency)
    all_results = []
    completed = [0]

    async def do_batch(batch_emails, batch_num):
        async with sem:
            try:
                async with session.post(
                    f"{api_url}/verify/batch",
                    json={"emails": batch_emails},
                    headers=headers,
                    timeout=aiohttp.ClientTimeout(total=600),
                ) as resp:
                    if resp.status == 200:
                        results = await resp.json()
                        results = results if isinstance(results, list) else results.get("results", [])
                        d = sum(1 for r in results if r.get("result") == "deliverable")
                        completed[0] += len(results)
                        click.echo(f"    [{label}] Batch {batch_num}/{len(batches)}: {d}/{len(results)} deliverable ({completed[0]}/{len(emails)})")
                        await asyncio.sleep(2)
                        return results
                    else:
                        click.echo(f"    [{label}] Batch {batch_num} error: {resp.status}")
                        return []
            except Exception as e:
                click.echo(f"    [{label}] Batch {batch_num} error: {e}")
                return []

    tasks = [do_batch(batch, i + 1) for i, batch in enumerate(batches)]
    for coro in asyncio.as_completed(tasks):
        all_results.extend(await coro)
    return all_results


# ── Squeeze Loop ────────────────────────────────────────────────────────────


async def squeeze_loop(
    session: aiohttp.ClientSession,
    email_results: dict[str, dict],
    api_url: str,
    headers: dict,
    max_iterations: int = MAX_SQUEEZE_ITERATIONS,
    cooldown: int = SQUEEZE_COOLDOWN,
) -> int:
    """Iteratively re-verify risky and unknown emails until dry. Returns total gained."""
    total_gained = 0
    zero_streak = 0

    for iteration in range(1, max_iterations + 1):
        risky = [e for e, r in email_results.items() if r.get("result") == "risky"]
        unknown = [e for e, r in email_results.items() if r.get("result") in ("unknown", "")]
        deliverable_count = sum(1 for r in email_results.values() if r.get("result") == "deliverable")
        catchall_count = sum(1 for r in email_results.values() if r.get("result") == "accept_all")

        if not risky and not unknown:
            click.echo(f"\n  No risky or unknown emails left. Done.")
            break

        click.echo(f"\n  SQUEEZE {iteration}/{max_iterations}: {deliverable_count} deliverable | {len(risky)} risky | {len(unknown)} unknown | {catchall_count} catch-all")

        gained = 0

        if risky:
            results = await _api_verify_batch(session, risky, api_url, headers, f"R{iteration}")
            for r in results:
                e = (r.get("email") or "").lower().strip()
                if not e:
                    continue
                old = email_results.get(e, {}).get("result", "")
                if r.get("result") == "deliverable" and old != "deliverable":
                    gained += 1
                if r.get("result") and r.get("result") != old:
                    email_results[e]["result"] = r["result"]

        if unknown:
            results = await _api_verify_batch(session, unknown, api_url, headers, f"U{iteration}", batch_size=100, concurrency=3)
            for r in results:
                e = (r.get("email") or "").lower().strip()
                if not e:
                    continue
                old = email_results.get(e, {}).get("result", "")
                if r.get("result") == "deliverable" and old != "deliverable":
                    gained += 1
                if r.get("result") and r.get("result") not in ("unknown", ""):
                    email_results[e]["result"] = r["result"]

        total_gained += gained
        click.echo(f"  Iteration {iteration}: +{gained} deliverable (cumulative: +{total_gained})")

        if gained == 0:
            zero_streak += 1
        else:
            zero_streak = 0

        if zero_streak >= 2:
            click.echo(f"  Two consecutive zero-gain iterations. Done.")
            break

        if iteration < max_iterations:
            click.echo(f"  Cooling {cooldown}s...")
            await asyncio.sleep(cooldown)

    return total_gained


# ── State Management ────────────────────────────────────────────────────────

STATE_FIELDS = [
    "first_name", "last_name", "full_name", "email", "company", "domain",
    "position", "website", "phone", "linkedin", "location", "profile_url",
    "source_file", "email_source", "result", "find_method", "find_confidence",
]


def save_state(contacts: list[dict], output_dir: Path):
    """Save processing state to CSV for resume capability."""
    output_dir.mkdir(parents=True, exist_ok=True)
    state_path = output_dir / "state.csv"
    with open(state_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=STATE_FIELDS, extrasaction="ignore")
        w.writeheader()
        for c in contacts:
            w.writerow(c)
    return state_path


def load_state(state_path: Path) -> list[dict]:
    """Load processing state from CSV."""
    with open(state_path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ── Export ──────────────────────────────────────────────────────────────────


def export_xlsx(contacts: list[dict], output_path: Path):
    """Export contacts with verification results to a formatted xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Color scheme
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )

    RESULT_FILL = {
        "deliverable": GREEN,
        "accept_all": YELLOW,
        "risky": YELLOW,
        "undeliverable": RED,
        "unknown": GRAY,
    }

    COLUMNS = [
        ("Full Name", "full_name", 22),
        ("First Name", "first_name", 14),
        ("Last Name", "last_name", 14),
        ("Email", "email", 30),
        ("Status", "result", 14),
        ("Email Source", "email_source", 16),
        ("Company", "company", 24),
        ("Position", "position", 24),
        ("Domain", "domain", 20),
        ("Website", "website", 24),
        ("Phone", "phone", 16),
        ("LinkedIn", "linkedin", 30),
        ("Location", "location", 18),
        ("Source File", "source_file", 22),
    ]

    def _write_sheet(ws, rows, name):
        ws.title = name
        # Header row
        for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, contact in enumerate(rows, 2):
            for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
                val = contact.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                if field == "result":
                    fill = RESULT_FILL.get(val, GRAY)
                    cell.fill = fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    # Segment contacts
    deliverable = [c for c in contacts if c.get("result") == "deliverable"]
    catchall = [c for c in contacts if c.get("result") == "accept_all"]
    risky = [c for c in contacts if c.get("result") == "risky"]
    unknown = [c for c in contacts if c.get("result") in ("unknown", "")]
    invalid = [c for c in contacts if c.get("result") == "undeliverable"]
    no_email = [c for c in contacts if not c.get("email")]

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 36
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 14

    summary_data = [
        ("KadenVerify Batch Results", "", ""),
        (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "", ""),
        ("", "", ""),
        ("Category", "Count", "Percent"),
        ("Total Contacts", len(contacts), "100%"),
        ("", "", ""),
        ("Deliverable (SMTP verified)", len(deliverable), f"{len(deliverable)/max(len(contacts),1)*100:.1f}%"),
        ("Catch-All (RISKY — may bounce)", len(catchall), f"{len(catchall)/max(len(contacts),1)*100:.1f}%"),
        ("Risky (other)", len(risky), f"{len(risky)/max(len(contacts),1)*100:.1f}%"),
        ("Unknown (UNVERIFIED — do NOT send)", len(unknown), f"{len(unknown)/max(len(contacts),1)*100:.1f}%"),
        ("Invalid (undeliverable)", len(invalid), f"{len(invalid)/max(len(contacts),1)*100:.1f}%"),
        ("No Email Found", len(no_email), f"{len(no_email)/max(len(contacts),1)*100:.1f}%"),
        ("", "", ""),
        ("SAFE TO SEND (upload this sheet)", len(deliverable), f"{len(deliverable)/max(len(contacts),1)*100:.1f}%"),
        ("", "", ""),
        ("WARNING: Only upload the 'Safe to Send' sheet to Instantly.", "", ""),
        ("Catch-all and unknown emails WILL bounce and damage sender reputation.", "", ""),
    ]

    for row_idx, (a, b, c) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=a)
        ws_summary.cell(row=row_idx, column=2, value=b)
        ws_summary.cell(row=row_idx, column=3, value=c)

    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.cell(row=4, column=1).font = Font(bold=True)
    ws_summary.cell(row=4, column=2).font = Font(bold=True)
    ws_summary.cell(row=4, column=3).font = Font(bold=True)
    ws_summary.cell(row=14, column=1).font = Font(bold=True, color="006100")
    ws_summary.cell(row=14, column=2).font = Font(bold=True, color="006100")
    ws_summary.cell(row=15, column=1).font = Font(bold=True, color="1F4E79")
    ws_summary.cell(row=15, column=2).font = Font(bold=True, color="1F4E79")

    # Data sheets
    _write_sheet(wb.create_sheet(), deliverable, "Safe to Send")
    _write_sheet(wb.create_sheet(), catchall, "Catch-All (RISKY)")
    _write_sheet(wb.create_sheet(), contacts, "All Contacts")
    if no_email:
        _write_sheet(wb.create_sheet(), no_email, "No Email")

    wb.save(output_path)
    return output_path


# ── Main Pipeline ───────────────────────────────────────────────────────────


async def run_pipeline(
    contacts: list[dict],
    output_dir: Path,
    api_url: str,
    api_key: str,
    squeeze: int,
):
    """Run the full find → verify → squeeze pipeline."""
    start = time.time()
    headers = {"Content-Type": "application/json", "X-API-Key": api_key}

    # Split contacts
    have_email = [c for c in contacts if c.get("email")]
    need_finding = [c for c in contacts if not c.get("email")]
    can_find = [c for c in need_finding if c.get("domain") and c.get("first_name") and c.get("last_name")]
    cant_find = [c for c in need_finding if not (c.get("domain") and c.get("first_name") and c.get("last_name"))]

    click.echo(f"\n  Total contacts: {len(contacts)}")
    click.echo(f"  Already have email: {len(have_email)}")
    click.echo(f"  Need finding: {len(need_finding)} ({len(can_find)} findable, {len(cant_find)} missing domain/name)")

    # Build contact lookup by (name, domain) for merging find results back
    # Use list values to handle duplicate (first, last, domain) tuples
    contact_lookup: dict[tuple, list[dict]] = {}
    for c in can_find:
        key = (c["first_name"].lower(), c["last_name"].lower(), c["domain"].lower())
        contact_lookup.setdefault(key, []).append(c)

    async with aiohttp.ClientSession() as session:
        # ── Phase 1: Find + Verify in parallel ──────────────────────────
        click.echo(f"\n{'='*60}")
        click.echo(f"  PHASE 1: Find emails + Verify existing (parallel)")
        click.echo(f"{'='*60}")

        emails_to_verify = list({c["email"] for c in have_email})

        find_coro = _api_find_batch(session, can_find, api_url, headers, "FIND")
        verify_coro = _api_verify_batch(session, emails_to_verify, api_url, headers, "VERIFY")

        find_results, verify_results = await asyncio.gather(find_coro, verify_coro)

        # Merge find results back into contacts
        found_emails = []
        for r in find_results:
            email = (r.get("email") or "").strip().lower()
            if not email or "@" not in email or float(r.get("confidence", 0) or 0) <= 0:
                continue
            found_emails.append(email)

            # Match back to all contacts with same (first, last, domain)
            first = (r.get("first_name") or "").lower()
            last = (r.get("last_name") or "").lower()
            domain = (r.get("domain") or "").lower()
            key = (first, last, domain)
            for contact in contact_lookup.get(key, []):
                contact["email"] = email
                contact["email_source"] = r.get("method", "found")
                contact["find_method"] = r.get("method", "")
                contact["find_confidence"] = str(r.get("confidence", 0))

        found_emails = list(set(found_emails))
        click.echo(f"\n  Found: {len(found_emails)}/{len(can_find)} unique emails")

        # Mark existing emails
        for c in have_email:
            if not c.get("email_source"):
                c["email_source"] = "original"

        # ── Phase 2: Verify newly found emails ──────────────────────────
        if found_emails:
            click.echo(f"\n  PHASE 2: Verify {len(found_emails)} newly found emails")
            found_verify = await _api_verify_batch(session, found_emails, api_url, headers, "FOUND-VERIFY")
            verify_results.extend(found_verify)

        # ── Build email result map ──────────────────────────────────────
        email_results = {}
        for r in verify_results:
            e = (r.get("email") or "").lower().strip()
            if e:
                email_results[e] = {"email": e, "result": r.get("result", "unknown")}

        # Assign results to contacts
        for c in contacts:
            e = c.get("email", "").lower().strip()
            if e and e in email_results:
                c["result"] = email_results[e]["result"]
            elif e:
                c["result"] = "unknown"
            else:
                c["result"] = ""

        # Save intermediate state
        save_state(contacts, output_dir)

        # ── Phase 3: Squeeze ────────────────────────────────────────────
        if squeeze > 0:
            click.echo(f"\n{'='*60}")
            click.echo(f"  PHASE 3: Squeeze (max {squeeze} iterations)")
            click.echo(f"{'='*60}")

            gained = await squeeze_loop(session, email_results, api_url, headers, max_iterations=squeeze)

            # Update contacts with squeezed results
            for c in contacts:
                e = c.get("email", "").lower().strip()
                if e and e in email_results:
                    c["result"] = email_results[e]["result"]

            click.echo(f"\n  Squeeze gained: +{gained} deliverable")

    # Save final state
    state_path = save_state(contacts, output_dir)

    elapsed = time.time() - start

    # Print final tally
    counts = Counter(c.get("result", "") for c in contacts)
    total_emails = sum(1 for c in contacts if c.get("email"))

    click.echo(f"\n{'='*60}")
    click.echo(f"  COMPLETE in {elapsed/60:.1f} min")
    click.echo(f"{'='*60}")
    click.echo(f"  Total contacts:  {len(contacts)}")
    click.echo(f"  Emails found:    {total_emails}")
    click.echo(f"  Deliverable:     {counts.get('deliverable', 0)}")
    click.echo(f"  Catch-all:       {counts.get('accept_all', 0)}")
    click.echo(f"  Risky:           {counts.get('risky', 0)}")
    click.echo(f"  Unknown:         {counts.get('unknown', 0)}")
    click.echo(f"  Invalid:         {counts.get('undeliverable', 0)}")
    click.echo(f"  No email:        {counts.get('', 0)}")
    click.echo(f"  SAFE TO SEND:    {counts.get('deliverable', 0)}  <-- only upload this to Instantly")
    click.echo(f"  Catch-all:       {counts.get('accept_all', 0)}  <-- DO NOT upload (will bounce)")
    click.echo(f"\n  State saved: {state_path}")

    return contacts


# ── CLI ─────────────────────────────────────────────────────────────────────


@click.group()
def main():
    """KadenVerify Batch Processor — find, verify, and export contact emails."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )


@main.command()
@click.argument("input_path", type=click.Path(exists=True))
@click.option("-o", "--output", "output_file", type=click.Path(), help="Output xlsx path")
@click.option("--output-dir", type=click.Path(), help="Working directory for state files")
@click.option("--api-url", default=DEFAULT_API_URL, help="KadenVerify API URL")
@click.option("--api-key", default=DEFAULT_API_KEY, help="API key")
@click.option("--squeeze", default=5, type=int, help="Max squeeze iterations (0 to skip)")
@click.option("--no-export", is_flag=True, help="Skip xlsx export")
def run(input_path, output_file, output_dir, api_url, api_key, squeeze, no_export):
    """Process contact files end-to-end: find → verify → squeeze → export.

    INPUT_PATH can be a single xlsx/csv file or a directory of files.
    """
    input_path = Path(input_path)

    # Resolve output directory
    if output_dir:
        out_dir = Path(output_dir)
    else:
        out_dir = Path(input_path.parent if input_path.is_file() else input_path) / "output"

    click.echo(f"{'='*60}")
    click.echo(f"  KadenVerify Batch Processor")
    click.echo(f"{'='*60}")
    click.echo(f"  Input:  {input_path}")
    click.echo(f"  Output: {out_dir}")
    click.echo(f"  API:    {api_url}")

    # Load contacts
    click.echo(f"\n  Loading contacts...")
    contacts = load_contacts(input_path)
    if not contacts:
        click.echo("No contacts found.")
        return

    # Run pipeline
    contacts = asyncio.run(run_pipeline(contacts, out_dir, api_url, api_key, squeeze))

    # Export
    if not no_export:
        if output_file:
            xlsx_path = Path(output_file)
        else:
            stamp = datetime.now().strftime("%Y-%m-%d")
            xlsx_path = out_dir / f"KadenVerify_Results_{stamp}.xlsx"

        click.echo(f"\n  Exporting to {xlsx_path}...")
        export_xlsx(contacts, xlsx_path)
        click.echo(f"  Done: {xlsx_path}")


@main.command()
@click.argument("state_file", type=click.Path(exists=True))
@click.option("-o", "--output", "output_file", type=click.Path(), required=True, help="Output xlsx path")
def export(state_file, output_file):
    """Re-export results from a saved state CSV to xlsx."""
    contacts = load_state(Path(state_file))
    if not contacts:
        click.echo("No contacts in state file.")
        return

    click.echo(f"  Loaded {len(contacts)} contacts from state")
    export_xlsx(contacts, Path(output_file))
    click.echo(f"  Exported to {output_file}")


@main.command()
@click.argument("state_file", type=click.Path(exists=True))
def stats(state_file):
    """Show statistics from a saved state CSV."""
    contacts = load_state(Path(state_file))
    if not contacts:
        click.echo("No contacts in state file.")
        return

    counts = Counter(c.get("result", "") for c in contacts)
    total = len(contacts)
    total_emails = sum(1 for c in contacts if c.get("email"))
    deliverable = counts.get("deliverable", 0)
    catchall = counts.get("accept_all", 0)

    click.echo(f"\n  Total contacts:  {total}")
    click.echo(f"  With email:      {total_emails}")
    click.echo(f"  Deliverable:     {deliverable} ({deliverable/max(total,1)*100:.1f}%)")
    click.echo(f"  Catch-all:       {catchall} ({catchall/max(total,1)*100:.1f}%)")
    click.echo(f"  Risky:           {counts.get('risky', 0)}")
    click.echo(f"  Unknown:         {counts.get('unknown', 0)}")
    click.echo(f"  Invalid:         {counts.get('undeliverable', 0)}")
    click.echo(f"  No email:        {counts.get('', 0)}")
    click.echo(f"\n  SAFE TO SEND:    {deliverable}")
    click.echo(f"  ALL USABLE:      {deliverable + catchall}")

    # Source file breakdown
    by_file = Counter(c.get("source_file", "unknown") for c in contacts)
    click.echo(f"\n  By source file:")
    for fname, count in by_file.most_common():
        file_deliverable = sum(1 for c in contacts if c.get("source_file") == fname and c.get("result") == "deliverable")
        click.echo(f"    {fname}: {count} contacts, {file_deliverable} deliverable")


if __name__ == "__main__":
    main()
